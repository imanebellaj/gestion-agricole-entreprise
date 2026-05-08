"""
Endpoints d'export — Excel et PDF pour marchés, projets, bénéficiaires.
"""
import io
from datetime import date

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from apps.marches.models import Marche
from apps.projets.models import Projet
from apps.acteurs.models import Beneficiaire


def _safe(v, default='—'):
    if v is None:
        return default
    s = str(v).strip()
    return s if s else default


# ─── Excel exports ───────────────────────────────────────────────────────────

class ExportMarchesExcel(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marchés"

        # Styles
        header_fill = PatternFill("solid", fgColor="052e16")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        alt_fill = PatternFill("solid", fgColor="F0FDF4")
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )
        center = Alignment(horizontal='center', vertical='center')

        # En-têtes
        headers = [
            'N° Marché', 'Année', 'Projet / Filière', 'Commune', 'Entreprise',
            'Montant engagé (DH)', 'Montant marché (DH)', 'Sup. potentielle (ha)',
            'Sup. plantée (ha)', 'Taux réalisation (%)', 'État',
            'Nb bénéficiaires', 'Dont jeunes', 'Dont femmes', 'Pénalités (DH)',
        ]

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        # Données
        marches = Marche.objects.select_related(
            'projet', 'projet__filiere', 'commune', 'entreprise'
        ).order_by('-annee', 'numero_marche')

        for row_idx, m in enumerate(marches, 2):
            sp = float(m.superficie_potentielle or 0)
            st = float(m.superficie_plantee or 0)
            taux = round(st / sp * 100, 1) if sp > 0 else 0

            row_data = [
                m.numero_marche,
                m.annee,
                f"{m.projet.intitule} ({m.projet.filiere.libelle})" if m.projet else '—',
                m.commune.libelle if m.commune else '—',
                m.entreprise.raison_sociale if m.entreprise else '—',
                float(m.montant_engage_dh or 0),
                float(m.montant_marche_dh or 0),
                float(m.superficie_potentielle or 0),
                float(m.superficie_plantee or 0),
                taux,
                m.get_etat_avancement_display(),
                m.nb_beneficiaires or 0,
                m.nb_beneficiaires_jeunes or 0,
                m.nb_beneficiaires_femmes or 0,
                float(m.penalite_retard_dh or 0),
            ]

            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                if fill:
                    cell.fill = fill
                if col_idx in (6, 7, 15):
                    cell.number_format = '#,##0.00'
                elif col_idx in (8, 9):
                    cell.number_format = '#,##0.00'
                elif col_idx == 10:
                    cell.number_format = '0.0"%"'

        # Largeurs colonnes
        widths = [15, 8, 35, 18, 25, 20, 20, 18, 16, 16, 14, 14, 12, 12, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # Onglet résumé
        ws2 = wb.create_sheet("Résumé")
        ws2['A1'] = 'Rapport Marchés — DPA Safi/Youssoufia'
        ws2['A1'].font = Font(bold=True, size=14, color="052e16")
        ws2['A3'] = f'Généré le: {date.today().strftime("%d/%m/%Y")}'
        ws2['A4'] = f'Total marchés: {marches.count()}'

        from django.db.models import Sum, Count
        totaux = marches.aggregate(
            mt=Sum('montant_engage_dh'), sp=Sum('superficie_plantee'), nb=Sum('nb_beneficiaires')
        )
        ws2['A5'] = f'Montant total engagé: {float(totaux["mt"] or 0):,.0f} DH'
        ws2['A6'] = f'Superficie totale plantée: {float(totaux["sp"] or 0):,.1f} ha'
        ws2['A7'] = f'Total bénéficiaires: {totaux["nb"] or 0}'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="marches_dpa_{date.today()}.xlsx"'
        return response


class ExportBeneficiairesExcel(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bénéficiaires"

        header_fill = PatternFill("solid", fgColor="1d4ed8")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        headers = ['CIN', 'Nom complet', 'Genre', 'Téléphone', 'Commune', 'Douar']
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        for row_idx, b in enumerate(Beneficiaire.objects.select_related('commune').order_by('nom_complet'), 2):
            for col_idx, val in enumerate([
                b.cin or '—', b.nom_complet,
                'Femme' if b.sexe == 'F' else 'Homme' if b.sexe == 'M' else '—',
                b.telephone or '—',
                b.commune.libelle if b.commune else '—',
                b.douar or '—',
            ], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border

        for i, w in enumerate([14, 30, 10, 14, 20, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="beneficiaires_dpa_{date.today()}.xlsx"'
        return response


class ExportProjetsPDF(APIView):
    """Export PDF d'un rapport synthétique des projets."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            topMargin=1.5*cm, bottomMargin=1.5*cm,
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            title="Rapport Projets DPA"
        )

        styles = getSampleStyleSheet()
        vert_dpa = colors.HexColor('#052e16')
        vert_clair = colors.HexColor('#dcfce7')

        title_style = ParagraphStyle(
            'title', fontSize=16, fontName='Helvetica-Bold',
            textColor=vert_dpa, alignment=TA_CENTER, spaceAfter=6
        )
        sub_style = ParagraphStyle(
            'sub', fontSize=10, fontName='Helvetica',
            textColor=colors.grey, alignment=TA_CENTER, spaceAfter=12
        )
        label_style = ParagraphStyle(
            'label', fontSize=9, fontName='Helvetica',
            textColor=colors.grey
        )

        elements = []

        # En-tête
        elements.append(Paragraph('DIRECTION PROVINCIALE DE L\'AGRICULTURE — SAFI / YOUSSOUFIA', title_style))
        elements.append(Paragraph(f'Rapport Marchés Agricoles — {date.today().strftime("%d %B %Y")}', sub_style))
        elements.append(HRFlowable(width="100%", thickness=2, color=vert_dpa, spaceAfter=12))

        # Tableau
        marches = list(Marche.objects.select_related(
            'projet__filiere', 'commune', 'entreprise'
        ).order_by('-annee', 'numero_marche')[:100])

        headers = [
            'N° Marché', 'An.', 'Filière', 'Commune', 'Entreprise',
            'Montant (MDH)', 'Sup. (ha)', 'Taux %', 'État'
        ]

        table_data = [headers]
        for m in marches:
            sp = float(m.superficie_potentielle or 0)
            st = float(m.superficie_plantee or 0)
            montant = float(m.montant_engage_dh or 0)
            taux = f"{st/sp*100:.0f}%" if sp > 0 else '—'
            etat_labels = {
                'en_cours': 'En cours', 'programme': 'Programmé',
                'receptionne': 'Réceptionné', 'cloture': 'Clôturé',
                'suspendu': 'Suspendu', 'resilie': 'Résilié',
            }
            table_data.append([
                m.numero_marche,
                str(m.annee),
                m.projet.filiere.libelle if m.projet and m.projet.filiere else '—',
                m.commune.libelle if m.commune else '—',
                (m.entreprise.raison_sociale[:20] if m.entreprise else '—'),
                f"{montant/1_000_000:.2f}" if montant > 0 else '—',
                f"{st:.1f}" if st > 0 else '—',
                taux,
                etat_labels.get(m.etat_avancement, m.etat_avancement),
            ])

        col_widths = [3.5*cm, 1.2*cm, 3*cm, 3*cm, 4*cm, 2.5*cm, 2*cm, 1.8*cm, 2.5*cm]

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), vert_dpa),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, vert_clair]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))

        # Totaux
        from django.db.models import Sum
        totaux = Marche.objects.aggregate(
            mt=Sum('montant_engage_dh'), sp=Sum('superficie_plantee'), nb=Sum('nb_beneficiaires')
        )
        elements.append(Paragraph(
            f"<b>Total marchés:</b> {Marche.objects.count()} | "
            f"<b>Montant total:</b> {float(totaux['mt'] or 0)/1_000_000:.2f} MDH | "
            f"<b>Sup. plantée:</b> {float(totaux['sp'] or 0):,.0f} ha | "
            f"<b>Bénéficiaires:</b> {totaux['nb'] or 0}",
            ParagraphStyle('footer', fontSize=9, fontName='Helvetica-Bold', textColor=vert_dpa)
        ))

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_marches_dpa_{date.today()}.pdf"'
        return response
