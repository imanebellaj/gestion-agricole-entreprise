"""
Commande d'import du fichier Excel MAJ mars 2026.xls
Usage: python manage.py import_excel --file "chemin/vers/MAJ mars 2026.xls"
"""
import re
import sys
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction

from apps.geo.models import Province, Commune
from apps.referentiels.models import Filiere, StatutProjet, Phase
from apps.acteurs.models import Entreprise
from apps.projets.models import Projet
from apps.marches.models import Marche, MarchePhase


# Mapping noms de feuilles Excel → libellé filière en base
FILIERES_MAP = {
    'OLIVIER': 'Olivier',
    'Olivier': 'Olivier',
    'FIGUIER': 'Figuier',
    'Figuier': 'Figuier',
    'CACTUS': 'Cactus Inerme',
    'Cactus inerme': 'Cactus Inerme',
    'CAPRIER': 'Câprier',
    'câprier': 'Câprier',
    'Câprier': 'Câprier',
    'CUMIN': 'Cumin',
    'Cumin': 'Cumin',
    'APICULTURE': 'Apiculture',
    'Apiculture': 'Apiculture',
    'VIANDES': 'Viandes Rouges',
    'Viandes rouges': 'Viandes Rouges',
}

# Communes par province (coordonnées GPS approximatives)
COMMUNES_GPS = {
    'Safi': {'lat': 32.2994, 'lon': -9.2372},
    'Youssoufia': {'lat': 32.2510, 'lon': -8.5294},
    'Chemaia': {'lat': 32.0833, 'lon': -8.9500},
    'Oulad Salmane': {'lat': 32.4000, 'lon': -9.0000},
    'Oulad Amrane': {'lat': 32.3500, 'lon': -8.8500},
    'Bouguedra': {'lat': 32.1500, 'lon': -9.0500},
    'Sidi Aissa': {'lat': 32.1000, 'lon': -9.1000},
    'Gharbia': {'lat': 32.2000, 'lon': -9.3000},
    'Had Hrara': {'lat': 32.0500, 'lon': -8.8000},
    'Laâtamna': {'lat': 32.3000, 'lon': -8.7000},
    'Lamrasla': {'lat': 32.1500, 'lon': -8.6000},
    'Ouled Sidi Yahya': {'lat': 32.2000, 'lon': -8.9000},
    'Sidi Chiker': {'lat': 32.1000, 'lon': -8.5000},
    'Nzala': {'lat': 32.4000, 'lon': -8.6000},
    'Tnine Ghiate': {'lat': 32.3000, 'lon': -9.1000},
}


def _to_decimal(val):
    if val is None or str(val).strip() in ('', '-', 'N/A', 'nan'):
        return None
    try:
        cleaned = str(val).replace(' ', '').replace(',', '.').replace('\xa0', '')
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _to_int(val):
    if val is None or str(val).strip() in ('', '-', 'N/A', 'nan'):
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _to_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val if isinstance(val, date) else val.date()
    s = str(val).strip()
    if s in ('', '-', 'N/A', 'nan'):
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _find_col(headers, *candidates):
    """Trouve la première colonne correspondant à l'une des options candidates."""
    norm = {h.strip().lower(): i for i, h in enumerate(headers) if h}
    for c in candidates:
        k = c.strip().lower()
        if k in norm:
            return norm[k]
    # Recherche partielle
    for c in candidates:
        k = c.strip().lower()
        for h_key, idx in norm.items():
            if k in h_key or h_key in k:
                return idx
    return None


class Command(BaseCommand):
    help = 'Importe les données du fichier Excel MAJ mars 2026'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, required=True, help='Chemin vers le fichier .xls ou .xlsx')
        parser.add_argument('--dry-run', action='store_true', help='Simulation sans écriture en base')
        parser.add_argument('--clear', action='store_true', help='Supprimer les marchés existants avant import')

    def handle(self, *args, **options):
        filepath = options['file']
        dry_run = options['dry_run']
        clear = options['clear']

        try:
            import openpyxl
            if filepath.endswith('.xls'):
                import xlrd
                wb = self._load_xls(filepath)
            else:
                wb = self._load_xlsx(filepath)
        except ImportError as e:
            self.stderr.write(f"[ERREUR] Bibliotheque manquante: {e}")
            sys.exit(1)
        except Exception as e:
            self.stderr.write(f"[ERREUR] Impossible d'ouvrir le fichier: {e}")
            sys.exit(1)

        stats = {'projets': 0, 'marches': 0, 'entreprises': 0, 'communes': 0, 'erreurs': 0}

        with transaction.atomic():
            if clear and not dry_run:
                Marche.objects.all().delete()
                self.stdout.write("[OK] Marches existants supprimes")

            for sheet_name, rows in wb.items():
                filiere_libelle = None
                for key, val in FILIERES_MAP.items():
                    if key.lower() in sheet_name.lower():
                        filiere_libelle = val
                        break

                if not filiere_libelle:
                    self.stdout.write(f"[SKIP] Feuille ignoree: {sheet_name}")
                    continue

                self.stdout.write(f"\n[SHEET] Traitement: {sheet_name} -> {filiere_libelle}")
                self._import_sheet(rows, filiere_libelle, dry_run, stats)

            if dry_run:
                transaction.set_rollback(True)
                self.stdout.write("\n[DRY-RUN] Aucune donnee sauvegardee")

        self.stdout.write(self.style.SUCCESS(
            f"\n[OK] Import termine:"
            f"\n   Projets    : {stats['projets']}"
            f"\n   Marches    : {stats['marches']}"
            f"\n   Entreprises: {stats['entreprises']}"
            f"\n   Communes   : {stats['communes']}"
            f"\n   Erreurs    : {stats['erreurs']}"
        ))

    def _load_xls(self, filepath):
        """Charge un fichier .xls et retourne un dict sheet_name → liste de rows (listes)."""
        import xlrd
        book = xlrd.open_workbook(filepath)
        result = {}
        for sheet in book.sheets():
            rows = []
            for r in range(sheet.nrows):
                row = []
                for c in range(sheet.ncols):
                    cell = sheet.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            dt = xlrd.xldate_as_datetime(cell.value, book.datemode)
                            row.append(dt.date())
                        except Exception:
                            row.append(cell.value)
                    else:
                        row.append(cell.value)
                rows.append(row)
            result[sheet.name] = rows
        return result

    def _load_xlsx(self, filepath):
        """Charge un fichier .xlsx et retourne un dict sheet_name → liste de rows."""
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        result = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            result[name] = rows
        return result

    def _import_sheet(self, rows, filiere_libelle, dry_run, stats):
        if len(rows) < 2:
            return

        # Trouver la ligne d'en-tête (première ligne non vide avec au moins 3 cellules)
        header_row = 0
        for i, row in enumerate(rows[:5]):
            non_empty = [c for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 3:
                header_row = i
                break

        headers = [str(c).strip() if c else '' for c in rows[header_row]]

        # Mapping des colonnes
        col = {
            'annee':        _find_col(headers, 'annee', 'année', 'an'),
            'numero':       _find_col(headers, 'n° marché', 'marché n°', 'num marche', 'num_marche', 'n°', 'numero'),
            'commune':      _find_col(headers, 'commune', 'lieu', 'localisation', 'douair', 'douar'),
            'entreprise':   _find_col(headers, 'entreprise', 'titulaire', 'prestataire', 'société'),
            'montant_eng':  _find_col(headers, 'montant engagé', 'montant engage', 'montant engagement', 'montant en dh', 'montant'),
            'montant_mrc':  _find_col(headers, 'montant marché', 'montant marche', 'montant du marché'),
            'sup_pot':      _find_col(headers, 'superficie potentielle', 'sup potentielle', 'sup. potentielle', 'superficie prog'),
            'sup_real':     _find_col(headers, 'superficie réalisée', 'sup realisee', 'sup. réalisée', 'realisée'),
            'sup_trav':     _find_col(headers, 'superficie travaillée', 'sup travaillee', 'sup. travaillée'),
            'sup_plant':    _find_col(headers, 'superficie plantée', 'sup plantee', 'plantée', 'sup. plantée'),
            'etat':         _find_col(headers, 'etat', 'état', 'état avancement', 'situation'),
            'penalite':     _find_col(headers, 'penalite', 'pénalité', 'penalites', 'pénalités'),
            'nb_benef':     _find_col(headers, 'nb bénéficiaires', 'nb beneficiaires', 'nombre bénéficiaires', 'nbre benef'),
            'nb_jeunes':    _find_col(headers, 'dont jeunes', 'nb jeunes', 'jeunes', 'jeune'),
            'nb_femmes':    _find_col(headers, 'dont femmes', 'nb femmes', 'femmes', 'femme'),
            'date_ods':     _find_col(headers, 'date ods', 'ods notification', 'ods', 'date notification'),
            'date_prev':    _find_col(headers, 'date prévue', 'reception prevue', 'date reception prevue', 'date prev'),
            'date_real':    _find_col(headers, 'date réelle', 'reception reelle', 'date reception reelle'),
            'phase':        _find_col(headers, 'phase', 'phase en cours'),
            'objet':        _find_col(headers, 'objet', 'intitulé', 'description', 'libellé'),
        }

        filiere = self._get_or_create_filiere(filiere_libelle)
        projet = self._get_or_create_projet(filiere_libelle, filiere, stats, dry_run)

        for i, row in enumerate(rows[header_row + 1:], start=header_row + 2):
            if all(c is None or str(c).strip() == '' for c in row):
                continue

            def get(key):
                idx = col.get(key)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    return v if v is not None else None
                return None

            # Numéro marché requis
            num = get('annee')
            if num is None:
                continue
            annee = _to_int(num)
            if annee is None or annee < 2000 or annee > 2030:
                continue

            numero_marche = str(get('numero') or '').strip()
            if not numero_marche:
                numero_marche = f"{filiere_libelle[:3].upper()}/{annee}/{i}"

            # Commune
            commune_libelle = str(get('commune') or '').strip()
            commune = self._get_or_create_commune(commune_libelle, stats, dry_run) if commune_libelle else None

            # Entreprise
            ent_nom = str(get('entreprise') or '').strip()
            entreprise = self._get_or_create_entreprise(ent_nom, stats, dry_run) if ent_nom else None

            # État
            etat_raw = str(get('etat') or '').lower().strip()
            etat = self._map_etat(etat_raw)

            # Objet
            objet = str(get('objet') or f"Travaux {filiere_libelle} - {commune_libelle}").strip()

            try:
                if not dry_run:
                    marche, created = Marche.objects.update_or_create(
                        numero_marche=numero_marche,
                        annee=annee,
                        defaults={
                            'projet': projet,
                            'commune': commune,
                            'entreprise': entreprise,
                            'objet': objet,
                            'montant_engage_dh': _to_decimal(get('montant_eng')),
                            'montant_marche_dh': _to_decimal(get('montant_mrc')),
                            'superficie_potentielle': _to_decimal(get('sup_pot')),
                            'superficie_realisee': _to_decimal(get('sup_real')),
                            'superficie_travaillee': _to_decimal(get('sup_trav')),
                            'superficie_plantee': _to_decimal(get('sup_plant')),
                            'penalite_retard_dh': _to_decimal(get('penalite')) or Decimal('0'),
                            'nb_beneficiaires': _to_int(get('nb_benef')),
                            'nb_beneficiaires_jeunes': _to_int(get('nb_jeunes')),
                            'nb_beneficiaires_femmes': _to_int(get('nb_femmes')),
                            'etat_avancement': etat,
                        }
                    )

                    # Phase si données disponibles
                    date_ods = _to_date(get('date_ods'))
                    date_prev = _to_date(get('date_prev'))
                    date_real = _to_date(get('date_real'))
                    phase_nom = str(get('phase') or '').strip()

                    if date_ods or date_prev:
                        phase_code = (phase_nom or 'plantation').lower().replace(' ', '_')[:30]
                        phase_obj, _ = Phase.objects.get_or_create(
                            code=phase_code,
                            defaults={'libelle': phase_nom or 'Plantation', 'ordre': 1}
                        )
                        MarchePhase.objects.update_or_create(
                            marche=marche,
                            phase=phase_obj,
                            defaults={
                                'date_ods_notification': date_ods,
                                'date_reception_prevue': date_prev,
                                'date_reception_reelle': date_real,
                                'superficie_plantee': _to_decimal(get('sup_plant')),
                            }
                        )

                stats['marches'] += 1

            except Exception as e:
                stats['erreurs'] += 1
                self.stderr.write(f"  [WARN] Ligne {i}: {e}")

        self.stdout.write(f"   -> {stats['marches']} marches traites")

    def _get_or_create_filiere(self, libelle):
        cat_map = {
            'Olivier': 'arboriculture', 'Figuier': 'arboriculture',
            'Cactus Inerme': 'arboriculture', 'Câprier': 'arboriculture',
            'Cumin': 'plantes_aromatiques', 'Apiculture': 'apiculture',
            'Viandes Rouges': 'elevage',
        }
        cat = cat_map.get(libelle, 'autre')
        filiere, _ = Filiere.objects.get_or_create(
            code=libelle[:30].upper().replace(' ', '_').replace('É', 'E').replace('Â', 'A'),
            defaults={'libelle': libelle, 'categorie': cat}
        )
        return filiere

    def _get_or_create_projet(self, filiere_libelle, filiere, stats, dry_run):
        province_safi, _ = Province.objects.get_or_create(
            code='SAF', defaults={'libelle': 'Safi', 'region': 'Marrakech-Safi'}
        )
        statut, _ = StatutProjet.objects.get_or_create(
            code='en_cours', defaults={'libelle': 'En cours'}
        )
        if not dry_run:
            projet = Projet.objects.filter(filiere=filiere).first()
            if not projet:
                projet = Projet.objects.create(
                    intitule=f"Programme {filiere_libelle} - DPA Safi/Youssoufia",
                    filiere=filiere,
                    province=province_safi,
                    statut=statut,
                )
                stats['projets'] += 1
            return projet
        return Projet.objects.filter(filiere=filiere).first()

    def _get_or_create_commune(self, libelle, stats, dry_run):
        if not libelle or libelle.lower() in ('', '-', 'n/a'):
            return None
        province, _ = Province.objects.get_or_create(
            code='SAF', defaults={'libelle': 'Safi'}
        )
        gps = COMMUNES_GPS.get(libelle, {})
        if not dry_run:
            commune, created = Commune.objects.get_or_create(
                libelle__iexact=libelle,
                defaults={
                    'libelle': libelle,
                    'province': province,
                    'latitude': gps.get('lat'),
                    'longitude': gps.get('lon'),
                }
            )
            if created:
                stats['communes'] += 1
            return commune
        return Commune.objects.filter(libelle__iexact=libelle).first()

    def _get_or_create_entreprise(self, nom, stats, dry_run):
        if not nom or nom.lower() in ('', '-', 'n/a'):
            return None
        if not dry_run:
            entreprise, created = Entreprise.objects.get_or_create(
                raison_sociale__iexact=nom,
                defaults={'raison_sociale': nom}
            )
            if created:
                stats['entreprises'] += 1
            return entreprise
        return Entreprise.objects.filter(raison_sociale__iexact=nom).first()

    def _map_etat(self, etat_raw):
        mapping = {
            'en cours': 'en_cours',
            'encours': 'en_cours',
            'programme': 'programme',
            'programmé': 'programme',
            'receptionne': 'receptionne',
            'réceptionné': 'receptionne',
            'cloture': 'cloture',
            'clôturé': 'cloture',
            'suspendu': 'suspendu',
            'resilie': 'resilie',
            'résilié': 'resilie',
            'cede': 'cede',
            'cédé': 'cede',
        }
        for k, v in mapping.items():
            if k in etat_raw:
                return v
        return 'en_cours'
